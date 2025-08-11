from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from ...models.dataset import Dataset
from ...models.dataset_feature import DatasetFeature
from ...models.dataset_product import DatasetProduct
from ...models.dataset_value import DatasetValue


class XlsxParser:
    def __init__(self, session: AsyncSession, dataset: Dataset):
        self.session = session
        self.dataset = dataset

    async def parse(self, bytes: BytesIO):
        worksheet = load_workbook(bytes).active
        if not worksheet:
            raise ValueError("No Worksheet in excel")

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return

        features = [
            DatasetFeature(dataset_id=self.dataset.id, name=str(header))
            for header in rows[0]
            if header is not None
        ]
        self.session.add_all(features)
        await self.session.flush()

        products = [
            DatasetProduct(dataset_id=self.dataset.id) for row in rows[1:] if any(row)
        ]
        self.session.add_all(products)
        await self.session.flush()

        for product, row in zip(products, (r for r in rows[1:] if any(r))):
            self.session.add_all(
                [
                    DatasetValue(
                        product_id=product.id,
                        feature_id=features[i].id,
                        value=str(cell),
                    )
                    for i, cell in enumerate(row[: len(features)])
                    if cell is not None
                ]
            )

        await self.session.commit()
